"""
Cnergee · Market Sales Briefing  –  app.py
==========================================
Column references aligned to:  Marketing Matrix 06-25 filled and cleaned-ba7e1fe9.xlsx
Row layout: HDR=5, SUB=6, DATA starts at 7, Country in col B (=2).

Run:  streamlit run app.py
"""

import io, re
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Cnergee · Market Sales Briefing",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
  .block-container{padding-top:1.1rem}
  h1{font-size:1.55rem!important;margin-bottom:.15rem!important}
  h2{font-size:1.25rem!important;margin-top:1.2rem!important}
  h3,h4{font-size:1.05rem!important;margin-top:.8rem!important}
  .stExpander summary p{font-size:.88rem}
  div[data-testid="metric-container"]>div:first-child{font-size:.7rem}
  div[data-testid="metric-container"]>div:last-child{font-size:1.15rem;font-weight:700}
</style>
""", unsafe_allow_html=True)

# ── SHEET LAYOUT ──────────────────────────────────────────────────────────────
HDR_ROW, SUB_ROW, DATA_START, COUNTRY_COL = 5, 6, 7, 2

# ── COLUMN MAP  (06-25 / ba7e1fe9 positions) ──────────────────────────────────
COLS = {
    # Infrastructure & Power
    "pwr_out_pct":   5,   # % of firms experiencing power outage (numeric)
    "trans_loss":    6,   # Grid transmission loss %  (numeric)
    "backup_u":      9,   # Power backup cost – Urban  (qual)
    "backup_r":     10,   # Power backup cost – Rural  (qual)
    "transport_q":  13,   # Transport infra quality score /5  (numeric)
    "access_u":     14,   # Ease of access to remote – Urban (qual, inverse fear)
    "access_r":     15,   # Ease of access to remote – Rural (qual, inverse fear)
    # Network
    "theft":        23,   # Network infra theft risk  (qual)
    "proj":         30,   # Projected 4G-5G growth to 2030 (text, e.g. "Rising (strong)")
    "inet_u":       31,   # Internet data charges – Urban (text)
    "inet_r":       32,   # Internet data charges – Rural (text)
    "aging_u":      34,   # Aging network infra – Urban (qual)
    "aging_r":      35,   # Aging network infra – Rural (qual)
    "ill_price":    36,   # ILL $/Mbps/month  (numeric)
    "fg_price":     37,   # 4G/5G $/Mbps/month  (numeric)
    "bb_price":     39,   # Broadband $/Mbps/month  (numeric)
    "bb_per100":    40,   # Fixed broadband per 100 people (numeric)
    "mob_per100":   41,   # Mobile subscriptions per 100  (numeric)
    "cloud":        42,   # Cloud adoption  (qual)
    # Human Resources
    "avail_l1":     43,   # Availability L1
    "avail_l2":     44,   # Availability L2
    "avail_l3":     45,   # Availability L3
    "sal_l1":       46,   # Salary L1 USD/yr  (numeric)
    "sal_l2":       47,   # Salary L2 USD/yr  (numeric)
    "sal_l3":       48,   # Salary L3 USD/yr  (numeric)
    "cyber_avail":  52,   # Cyber security experts availability (qual)
    "unemployment": 54,   # Unemployment ratio %  (numeric or text "4.2%")
    "skill":        55,   # Skill gap  (qual)
    # Currency / GDP
    "curr_usd":     57,   # Currency vs USD
    "curr_inr":     58,   # Currency vs INR
    "gdp":          59,   # GDP USD  (text, e.g. "$3.91tn")
    "growth":       62,   # GDP growth rate %  (numeric or "6.5%")
    "tech_vision":  63,   # Tech vision for 2030  (text)
    "recession":    64,   # Recession risk  (qual)
    "curfluct":     65,   # Currency fluctuation  (qual)
    "ease_import":  68,   # Ease of import  (qual, inverse fear)
    "crit_infra":   70,   # Critical infra dependency /10  (numeric)
    "import_duty":  74,   # Import duties %  (numeric or "4.6%")
    # Downtime
    "dt_cult":      75,   # Downtime – cultural sensitivity  (qual)
    "dt_rep":       76,   # Downtime – reputation loss  (qual)
    "dt_biz":       77,   # Downtime – business loss  (qual)
    "downtime_cost":78,   # Cost of downtime USD/hr  (numeric)
    # Telecom / ISP
    "isp_comp":     79,   # ISP competition  (qual)
    "local_isp":    80,   # Local ISP names  (text)
    "top_telecom":  81,   # Top telecom operator  (text)
    "telecom_trend":82,   # Telecom revenue trend  (text)
    "arpu":         83,   # Current ARPU USD/month  (numeric)
    "arpu_2030":    84,   # Projected ARPU 2030  (numeric)
    "churn":        85,   # Telecom churn %/yr  (numeric)
    # TCO
    "customer_tco": 87,   # Customer 3-yr TCO  (numeric)
    "cnergee_cost": 88,   # Cnergee 3-yr cost  (numeric)
    "savings_tco":  89,   # 3-yr savings  (numeric)
    "msp_opp":      90,   # Managed service opportunity  (text)
    "auto_spend":   91,   # Spendings on automation  (qual)
    "auto_proj":    92,   # Projected automation spend  (text)
    "cyber_gdp":    93,   # Cyber security spend % of GDP  (numeric)
    # Cyber
    "attack_pct":   95,   # % of global cyber attacks  (numeric)
    "info_warfare": 96,   # Information warfare threat  (qual)
    "attack_2030":  97,   # Cyber attacks forecast 2030  (text)
    "data_sov":     98,   # Data sovereignty laws  (text)
    # Geo-political
    "embargo":      99,
    "geo_loc":     100,
    "war":         101,   # Eminent war / conflict threat  (qual)
    # TAM / SAM / SOM
    "tam_wifi":    102,
    "tam_sdwan":   103,
    "tam_ngfw":    104,
    "total_tam":   105,
    "total_sam":   106,
    "total_som":   107,
    # Demographics
    "literacy":    108,
    "pop_u":       110,   # Population Urban %
    "pop_r":       111,   # Population Rural %
    # Psychology
    "psych":       112,
    "brand":       113,
    "colors":      114,
    "belief":      115,
    # Taglines
    "tag_neg":     116,   # Value challenge line  (e.g. "Not Brand. Value.")
    "tag_pos":     117,   # Positive brand line
}

# ── SCORING HELPERS ───────────────────────────────────────────────────────────
def score(v):
    """Qualitative → 1-5 (higher = more intense / more of the quality described)."""
    if v is None: return None
    s = str(v).strip().lower()
    if "very high"      in s: return 5
    if "high"           in s and "very" not in s: return 4
    if "moderate-high"  in s: return 4
    if "moderate"       in s or "mixed" in s or "limited" in s: return 3
    if "very low"       in s: return 1
    if "low"            in s and "very" not in s: return 2
    # Infrastructure / adoption adjectives
    if any(w in s for w in ["significantly aging","heavily aging","very aging"]): return 5
    if any(w in s for w in ["aging","aging network","old"]): return 4
    if "modern" in s or "new" in s or "not aging" in s: return 2
    if any(w in s for w in ["widely adopted","very widely"]): return 5
    if "widely" in s and "not" not in s: return 4
    if "not widely" in s or "not adopted" in s: return 1
    # Rising / declining
    if "rising" in s and "strong" in s: return 5
    if "rising" in s and "moderate" in s: return 4
    if "rising" in s: return 4
    if "flat" in s or "stable" in s or "maturing" in s: return 2
    if "declining" in s: return 1
    # Access / import quality (positive = low fear via neg_score)
    if "very good" in s or "excellent" in s or "very easy" in s or "easy" in s: return 1
    if "good" in s and "very" not in s: return 2
    if "poor" in s and "very" not in s: return 4
    if "very poor" in s: return 5
    if "challenging" in s and "very" not in s: return 4
    if "very challenging" in s: return 5
    return None

def neg_score(v):
    """Inverse: low-quality text → high fear score.  'Poor' access → 5."""
    s_val = score(v)
    if s_val is None: return None
    return 6 - s_val   # 5↔1, 4↔2, 3↔3

def sev_label(s):
    if s >= 4.5: return ("Very High", "vh")
    if s >= 3.5: return ("High",      "h")
    return              ("Moderate",  "m")

SEV_COLOR = {"vh": "#ef4444", "h": "#f97316", "m": "#f59e0b"}

def parse_float(v):
    """Strip %, $, commas → float.  Returns None on failure."""
    if v in (None, ""): return None
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(v).split()[0]))
    except: return None

def parse_money(v):
    if v in (None, ""): return None
    try:
        n = float(re.sub(r"[^0-9.]", "", str(v)))
        return n if n > 0 else None
    except: return None

def fmt_money(n):
    if n is None: return "—"
    if n >= 1_000_000: return f"${n/1e6:.2f}M"
    if n >= 1_000:     return f"${n/1e3:.0f}K"
    return f"${n:.0f}"

# ── LOAD WORKBOOK ─────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Reading matrix…")
def load_sheet(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(HDR_ROW, c).value
        s = ws.cell(SUB_ROW, c).value
        lbl = " – ".join(str(x).strip() for x in [h, s] if x not in (None, ""))
        headers[c] = lbl if lbl else get_column_letter(c)
    rows = {}
    for r in range(DATA_START, ws.max_row + 1):
        nm = ws.cell(r, COUNTRY_COL).value
        if not nm: continue
        nm = str(nm).strip()
        if nm in rows: continue          # skip duplicates (e.g. Nigeria row 46)
        rows[nm] = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
    return headers, rows

# ── FEARS ─────────────────────────────────────────────────────────────────────
def fears(d):
    """
    Returns list of (score_float, name, severity_label, severity_class, pitch_line).
    Only items at Moderate (≥3) or above are returned.
    """
    out = []

    def add(s, nm, ln):
        if s is None or s < 3: return
        lbl, cls = sev_label(s)
        out.append((s, nm, lbl, cls, ln))

    def add_low(s, nm, ln):
        """For 'availability' cols: LOW score means scarce → fear."""
        if s is None or s > 2: return
        out.append((4.0, nm, "High", "h", ln))

    # ── NUMERIC FEARS ──

    # Col 5 – Power outage %
    po = parse_float(d.get(COLS["pwr_out_pct"]))
    if po and po >= 20:
        s = 5 if po >= 50 else (4 if po >= 35 else 3)
        add(s, f"Power Outage – {po:.0f}% of Firms Affected",
            f"{po:.0f}% of businesses lose power at least once a year. Every outage is lost revenue. "
            "Position 4G/5G failover as the always-on business continuity layer — power-independent and zero-touch.")

    # Col 6 – Grid transmission loss
    tl = parse_float(d.get(COLS["trans_loss"]))
    if tl and tl >= 15:
        s = 5 if tl >= 30 else (4 if tl >= 20 else 3)
        add(s, f"Grid Transmission Loss ({tl:.1f}%)",
            f"{tl:.1f}% of generated electricity never reaches the customer — the grid is fundamentally unreliable. "
            "Sell resilience as a core infrastructure requirement, not an optional add-on.")

    # Col 13 – Transport quality (low score = poor = high fear)
    tq = parse_float(d.get(COLS["transport_q"]))
    if tq and tq <= 3.0:
        s = 5 if tq <= 2.3 else (4 if tq <= 2.7 else 3)
        add(s, f"Poor Transport Infrastructure ({tq}/5)",
            f"Logistics quality rated {tq}/5 — truck-rolls, hardware delivery, and on-site engineers are slow and expensive. "
            "Zero-touch provisioning removes the physical visit from the equation entirely.")

    # Col 40 – Fixed broadband per 100 (low = wired desert = fear/opportunity)
    bb = parse_float(d.get(COLS["bb_per100"]))
    if bb is not None and bb <= 10:
        s = 5 if bb <= 3 else (4 if bb <= 7 else 3)
        add(s, f"Near-Zero Wired Internet ({bb:.1f}/100 people)",
            f"Only {bb:.1f} fixed broadband connections per 100 people — enterprises are trapped on expensive leased lines "
            "or patchy mobile. Cnergee Bonded Broadband aggregates 4 mobile links into a single enterprise-grade pipe: faster, cheaper, resilient.")

    # Col 54 – Unemployment
    ue = parse_float(d.get(COLS["unemployment"]))
    if ue and ue >= 15:
        s = 5 if ue >= 25 else (4 if ue >= 20 else 3)
        add(s, f"High Unemployment ({ue:.1f}%)",
            f"{ue:.1f}% unemployment signals economic stress — decision-makers are defensive on capex. "
            "Lead with cost reduction: show the TCO delta, not the feature list.")

    # Col 70 – Critical infrastructure dependency /10
    ci = parse_float(d.get(COLS["crit_infra"]))
    if ci and ci >= 7:
        s = 5 if ci >= 10 else (4 if ci >= 8 else 3)
        add(s, f"Critical Infrastructure Exposure ({ci:.0f}/10)",
            f"Rated {ci:.0f}/10 — a network outage or breach here has national-scale consequences. "
            "Enterprise-grade security and guaranteed uptime are a compliance requirement, not a preference.")

    # Col 74 – Import duties %
    duty = parse_float(d.get(COLS["import_duty"]))
    if duty and duty >= 5:
        s = 5 if duty >= 9 else (4 if duty >= 7 else 3)
        add(s, f"Import Duties ({duty:.1f}%) Add to Hardware Cost",
            f"{duty:.1f}% import duties inflate every hardware procurement. "
            "Cnergee's hardware-light, cloud-managed platform minimises physical equipment on site — less to import, lower duty exposure, faster deployment.")

    # Col 78 – Downtime cost USD/hr
    dtc = parse_money(d.get(COLS["downtime_cost"]))
    if dtc and dtc >= 2000:
        s = 5 if dtc >= 5000 else (4 if dtc >= 3000 else 3)
        add(s, f"Downtime Costs ${dtc:,.0f}/hr",
            f"Every hour of network failure costs this business ${dtc:,.0f}. Self-managed infrastructure averages ~100 outage hours/3 years. "
            f"Cnergee guarantees 99.5% uptime (≤44 hrs/yr). Run the numbers — it makes itself.")

    # Col 95 – % of global cyber attacks
    ap = parse_float(d.get(COLS["attack_pct"]))
    if ap and ap >= 0.3:
        s = 5 if ap >= 5 else (4 if ap >= 1 else 3)
        add(s, f"{ap:.2f}% of Global Cyber Attacks Aimed Here",
            f"This country accounts for {ap:.2f}% of all global cyber incidents — it is a named, targeted market. "
            "NGFW + centralised threat policy enforcement is the minimum viable security posture, not a premium option.")

    # Col 47 – L2 salary (high salary = expensive to self-manage)
    sal_l2 = parse_float(d.get(COLS["sal_l2"]))
    if sal_l2 and sal_l2 >= 20_000:
        s = 5 if sal_l2 >= 100_000 else (4 if sal_l2 >= 60_000 else 3)
        add(s, f"L2 Engineer Costs ${sal_l2:,.0f}/yr",
            f"Mid-level network engineers cost ${sal_l2:,.0f}/yr here — every manual maintenance task burns expensive opex. "
            "Automation reduces the headcount needed without reducing oversight quality.")

    # Col 48 – L3 salary
    sal_l3 = parse_float(d.get(COLS["sal_l3"]))
    if sal_l3 and sal_l3 >= 40_000:
        s = 5 if sal_l3 >= 150_000 else (4 if sal_l3 >= 100_000 else 3)
        add(s, f"L3 Specialist Costs ${sal_l3:,.0f}/yr",
            f"Senior network/security architects at ${sal_l3:,.0f}/yr are too expensive to keep on staff for routine ops. "
            "Cnergee's AI self-healing eliminates the L3 dependency from day-to-day network management.")

    # ── QUALITATIVE FEARS ──

    add(score(d.get(COLS["dt_rep"])),
        "Downtime – Reputation Loss",
        "An outage here costs trust that is hard to rebuild. Uptime is brand protection — position it that way.")

    add(score(d.get(COLS["dt_biz"])),
        "Downtime – Direct Business Loss",
        "Downtime bleeds revenue. Quantify their hourly cost, then make it go away.")

    add(score(d.get(COLS["dt_cult"])),
        "Downtime – Cultural Sensitivity",
        "Service interruptions carry outsized weight here. Reliability is a relationship issue, not just technical.")

    add(score(d.get(COLS["theft"])),
        "Network Infrastructure Theft",
        "Cable, fiber, and battery theft drives repeat failures and expensive call-outs. "
        "Hardware-light managed networking cuts the physical footprint that gets stolen.")

    aging = max(score(d.get(COLS["aging_u"])) or 0, score(d.get(COLS["aging_r"])) or 0) or None
    add(aging,
        "Aging Network Infrastructure",
        "Legacy kit needs constant patching and breaks without warning. "
        "Leapfrog it — modern managed infrastructure replaces the whole maintenance cycle.")

    add(score(d.get(COLS["skill"])),
        "Local Skill Gap",
        "Not enough skilled engineers to run complex infrastructure. "
        "A managed/automated solution closes the gap without adding permanent headcount.")

    add(score(d.get(COLS["recession"])),
        "Recession Risk",
        "Budgets are defensive — lead with cost reduction, not new capabilities. The TCO comparison carries the conversation.")

    add(score(d.get(COLS["curfluct"])),
        "Currency Fluctuation Risk",
        "Unstable FX erodes infrastructure budgets. Cnergee's fixed 3-year pricing locks in predictable total cost of ownership.")

    backup = max(score(d.get(COLS["backup_u"])) or 0, score(d.get(COLS["backup_r"])) or 0) or None
    add(backup,
        "High Power Backup Costs",
        "Keeping kit powered under unreliable grid conditions is expensive. "
        "Every hardware unit removed from site cuts the backup cost.")

    add(score(d.get(COLS["info_warfare"])),
        "Information Warfare Threat",
        "State-level information operations target enterprise networks. "
        "The threat is not just criminal — centralised, always-updated NGFW is essential.")

    add(score(d.get(COLS["war"])),
        "Conflict / War Proximity",
        "Active conflict nearby disrupts supply chains and expands the attack surface. "
        "Build for resilience and continuity, not just performance.")

    # Ease of access (neg_score: 'Poor' access → high fear)
    access = max(neg_score(d.get(COLS["access_u"])) or 0,
                 neg_score(d.get(COLS["access_r"])) or 0) or None
    add(access,
        "Hard Physical Access to Remote Sites",
        "Difficult terrain = expensive truck-rolls, slow incident response, high on-site cost. "
        "Zero-touch provisioning and remote management remove the physical visit dependency entirely.")

    # Ease of import (neg_score: 'Challenging' → high fear)
    ei = neg_score(d.get(COLS["ease_import"]))
    add(ei,
        "Difficult Import Environment",
        "Hardware import is slow and complex — every device replacement is a delay. "
        "Hardware-light, cloud-managed architecture reduces physical import dependency dramatically.")

    # Cyber attacks forecast 2030
    fc = str(d.get(COLS["attack_2030"]) or "").lower()
    if "rising" in fc:
        s = 5 if "strong" in fc else 4
        add(s, "Cyber Attack Volume Rising to 2030",
            "The threat environment gets worse, not better, through 2030. "
            "Selling NGFW + managed security now means selling into an expanding problem — urgency is on your side.")

    # Data sovereignty
    ds = str(d.get(COLS["data_sov"]) or "").lower()
    if "restrictive" in ds:
        s = 5 if "highly" in ds else (4 if "very" in ds else (3 if "moderate" in ds else 4))
        add(s, "Data Sovereignty / Compliance Burden",
            "Local data laws require compliance infrastructure — audit logs, policy enforcement, data residency. "
            "Cnergee NMS with centralised policy management and full audit trail aligns to these requirements automatically.")

    # Availability scarce (inverse: LOW availability score → high fear)
    av2 = score(d.get(COLS["avail_l2"]))
    av3 = score(d.get(COLS["avail_l3"]))
    if av2 and av3:
        av_worst = min(av2, av3)
        add_low(av_worst, "L2/L3 Engineers Hard to Find",
                "Thin local talent pool for network ops — a remote-managed model fills the capability gap they simply can't hire for.")

    add_low(score(d.get(COLS["cyber_avail"])),
            "Cyber Security Experts Scarce",
            "Few local defenders available to hire — managed security covers what they cannot staff internally.")

    out.sort(key=lambda x: x[0], reverse=True)
    return [(nm, lbl, cls, ln) for _s, nm, lbl, cls, ln in out]

# ── GREEDS ────────────────────────────────────────────────────────────────────
GREED_CATS = ["🟢 MARKET PULL", "📶 CONNECTIVITY WAVE", "💰 ACTIVE BUDGET", "🚪 OPEN DOOR"]
GREED_COLS = {
    "🟢 MARKET PULL":      "#16a34a",
    "📶 CONNECTIVITY WAVE":"#0891b2",
    "💰 ACTIVE BUDGET":    "#7c3aed",
    "🚪 OPEN DOOR":        "#d97706",
}

def greeds(d):
    """Returns dict: category → sorted list of (score, name, pitch_line)."""
    cats = {c: [] for c in GREED_CATS}

    def add(cat, s, nm, ln):
        cats[cat].append((s, nm, ln))

    # ── 🟢 MARKET PULL ──

    g = parse_float(d.get(COLS["growth"]))
    if g is not None:
        if g >= 7:
            add("🟢 MARKET PULL", 5, f"High-Velocity Economy ({g}% GDP Growth)",
                "Businesses are expanding fast and need infrastructure now. Urgency is already on your side.")
        elif g >= 4:
            add("🟢 MARKET PULL", 4, f"Fast-Growing Economy ({g}% GDP Growth)",
                "Strong growth creates new sites, new headcount, new network demand. Timing favors you.")
        elif g >= 2:
            add("🟢 MARKET PULL", 3, f"Stable Economy ({g}% Growth)",
                "Steady conditions — compete on efficiency and predictable total cost, not urgency.")

    gdp = d.get(COLS["gdp"])
    if gdp and str(gdp).strip() not in ("", "None"):
        add("🟢 MARKET PULL", 3, f"Market Size: {gdp}",
            "Absolute GDP sets the budget ceiling — a larger economy has a deeper IT spend pool.")

    tv = d.get(COLS["tech_vision"])
    if tv and str(tv).strip() not in ("", "None"):
        add("🟢 MARKET PULL", 4, "Government Technology Vision 2030",
            f"{str(tv)[:150]}. Government mandate creates enterprise urgency to align infrastructure with national direction.")

    tt = str(d.get(COLS["telecom_trend"]) or "").lower()
    if "growing" in tt or "rising" in tt:
        add("🟢 MARKET PULL", 4, "Telecom Sector Growing",
            "Telecom revenue is expanding — the market is active, not contracting. Buyers are investing, not freezing.")

    # ── 📶 CONNECTIVITY WAVE ──

    proj = str(d.get(COLS["proj"]) or "").lower()
    if "rising" in proj:
        s = 5 if "strong" in proj else 4
        add("📶 CONNECTIVITY WAVE", s, "Heavy 4G/5G Investment Pipeline",
            "Strong planned investment in mobile connectivity — position Cnergee as the efficient enterprise on-ramp to that wave.")

    mob = parse_float(d.get(COLS["mob_per100"]))
    if mob and mob >= 100:
        add("📶 CONNECTIVITY WAVE", 5 if mob >= 150 else 4,
            f"Mobile-First Country ({mob:.0f} SIMs/100 people)",
            f"{mob:.0f} mobile subscriptions per 100 people — bonded 4G/5G is not a workaround here, it's the enterprise-grade primary path.")

    bb100 = parse_float(d.get(COLS["bb_per100"]))
    if bb100 is not None and bb100 <= 10:
        add("📶 CONNECTIVITY WAVE", 4,
            f"Connectivity Gap = Our Opportunity ({bb100:.1f} Broadband/100)",
            "Almost no wired internet — Cnergee Bonded Broadband aggregates mobile links into a single resilient enterprise pipe. We ARE the wired alternative.")

    cl = score(d.get(COLS["cloud"]))
    if cl and cl >= 3:
        add("📶 CONNECTIVITY WAVE", cl, "Strong Cloud Adoption",
            "Cloud-ready buyers need a secure, reliable network layer to support it. Slot in as that layer — we're pre-positioned.")

    arpu = parse_float(d.get(COLS["arpu"]))
    arpu30 = parse_float(d.get(COLS["arpu_2030"]))
    if arpu and arpu30 and arpu30 > arpu:
        gpct = (arpu30 - arpu) / arpu * 100
        add("📶 CONNECTIVITY WAVE", 4 if gpct >= 30 else 3,
            f"ARPU Rising ${arpu:.0f}→${arpu30:.0f}/mo by 2030",
            f"Telecom ARPU projected to grow {gpct:.0f}% by 2030 — buyers are already spending more on connectivity, signalling higher willingness to pay for quality.")

    # ── 💰 ACTIVE BUDGET ──

    cg = parse_float(d.get(COLS["cyber_gdp"]))
    if cg and cg >= 0.8:
        add("💰 ACTIVE BUDGET", 5 if cg >= 1.2 else 4,
            f"Cyber Security Budget {cg:.1f}% of GDP",
            f"{cg:.1f}% of GDP already allocated to cyber security — the budget exists. The question is whether it's being spent on the right things.")

    asp = score(d.get(COLS["auto_spend"]))
    if asp and asp >= 3:
        add("💰 ACTIVE BUDGET", asp, "Active Automation Spend",
            "Already investing in automation — buyers understand managed infrastructure and have approved budgets for it.")

    apj = str(d.get(COLS["auto_proj"]) or "").lower()
    if "rising" in apj:
        add("💰 ACTIVE BUDGET", 4 if "strong" in apj else 3, "Automation Budget Growing",
            "Automation spend is growing year on year. The pipeline expands automatically — get in now.")

    msp = str(d.get(COLS["msp_opp"]) or "")
    if msp.strip() and msp.strip().lower() not in ("none", "—", "-"):
        add("💰 ACTIVE BUDGET", 3, "Active Managed Service Market",
            f"{msp[:180]}. Understanding how buyers purchase is half the battle — align the commercial model to what already works here.")

    # ── 🚪 OPEN DOOR ──

    churn = parse_float(d.get(COLS["churn"]))
    if churn and churn >= 22:
        add("🚪 OPEN DOOR", 5 if churn >= 28 else 4,
            f"Buyers Already Leaving Their Providers ({churn:.0f}% Churn/yr)",
            f"{churn:.0f}% annual churn — customers are dissatisfied and switching. The door is open. Lead with the switch conversation.")

    isp = score(d.get(COLS["isp_comp"]))
    if isp and isp >= 3:
        add("🚪 OPEN DOOR", isp, "Competitive ISP Market",
            "High ISP competition means buyers know they have options and are not loyal by inertia. Value-based conversations land easily.")

    result = {c: sorted(items, key=lambda x: x[0], reverse=True)
              for c, items in cats.items() if items}
    return result

# ── COST ANIMATION HTML ───────────────────────────────────────────────────────
def anim_html(cust: float, cnrg: float) -> str:
    return f"""<style>
  body{{margin:0;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif}}
  .wrap{{background:linear-gradient(90deg,#241410,#1a1410);border:1px solid #5a3a1a;
         border-radius:12px;padding:16px 18px}}
  .row{{display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap}}
  .lbl{{color:#d89a5a;font-size:11px;text-transform:uppercase;letter-spacing:.6px;font-weight:700}}
  .num{{font-size:32px;font-weight:800;color:#fbbf24;margin:4px 0}}
  .cap{{color:#9a8a72;font-size:11px}}
  .barwrap{{background:#2a2018;border-radius:8px;height:16px;overflow:hidden;margin:12px 0 4px}}
  .bar{{height:100%;width:100%;background:linear-gradient(90deg,#f97316,#fbbf24);
        transition:width 1.6s cubic-bezier(.2,.7,.2,1)}}
  .btn{{background:linear-gradient(90deg,#22d3ee,#2dd4bf);color:#04222a;border:none;
        border-radius:30px;padding:11px 22px;font-weight:800;font-size:14px;
        cursor:pointer;font-family:inherit;white-space:nowrap}}
  .sav{{color:#86efac;font-weight:700;font-size:14px;margin-top:8px;
        opacity:0;transition:opacity .6s}}
  .on .num{{color:#86efac}}
  .on .bar{{background:linear-gradient(90deg,#16a34a,#2dd4bf)}}
</style>
<div class="wrap" id="w">
  <div class="row">
    <div>
      <div class="lbl" id="lab">You're bleeding this now &middot; 3-yr self-managed</div>
      <div class="num" id="n">$0</div>
      <div class="cap" id="cap">Customer total cost of ownership &middot; per 100-seat office (est.)</div>
    </div>
    <button class="btn" id="b">&#9889; Switch on Cnergee</button>
  </div>
  <div class="barwrap"><div class="bar" id="bar"></div></div>
  <div class="sav" id="sav"></div>
</div>
<script>
const CUST={int(cust)}, CNRG={int(cnrg)};
const n=document.getElementById('n'), bar=document.getElementById('bar'),
      b=document.getElementById('b'), sav=document.getElementById('sav'),
      w=document.getElementById('w'), lab=document.getElementById('lab'),
      cap=document.getElementById('cap');
function fmt(x){{x=Math.round(x);
  if(x>=1e6)return'$'+(x/1e6).toFixed(2)+'M';
  if(x>=1e3)return'$'+Math.round(x/1e3)+'K';return'$'+x;}}
n.textContent=fmt(CUST);
let done=false;
b.onclick=function(){{
  if(done){{w.classList.remove('on');bar.style.width='100%';sav.style.opacity=0;
    lab.textContent="You're bleeding this now · 3-yr self-managed";
    cap.textContent="Customer total cost of ownership · per 100-seat office (est.)";
    b.textContent="⚡ Switch on Cnergee";n.textContent=fmt(CUST);done=false;return;}}
  w.classList.add('on');
  bar.style.width=Math.min(100,Math.max(4,(CNRG/CUST*100)))+'%';
  b.textContent="✓ Cnergee is running it";
  lab.textContent="With Cnergee · 3-yr managed";
  cap.textContent="Cnergee total cost · per 100-seat office (est.)";
  const start=performance.now(), dur=1600;
  function step(t){{let p=Math.min((t-start)/dur,1);p=1-Math.pow(1-p,3);
    n.textContent=fmt(CUST+(CNRG-CUST)*p);
    if(p<1)requestAnimationFrame(step);
    else{{sav.textContent="You save "+fmt(CUST-CNRG)+" over 3 years";
          sav.style.opacity=1;done=true;}}}}
  requestAnimationFrame(step);
}};
</script>"""

# ── HTML CARD HELPERS ─────────────────────────────────────────────────────────
def fear_card(nm, lbl, cls, ln):
    col = SEV_COLOR.get(cls, "#f59e0b")
    return (
        f"<div style='border-left:3px solid {col};background:#1b222b;border-radius:8px;"
        f"padding:9px 12px;margin-bottom:9px'>"
        f"<b style='color:#e2e8f0'>{nm}</b> "
        f"<span style='background:{col};color:#0f1923;font-size:10px;font-weight:700;"
        f"padding:1px 7px;border-radius:10px'>{lbl}</span><br>"
        f"<span style='color:#94a3b8;font-size:12.5px;line-height:1.5'>{ln}</span></div>"
    )

def greed_card(nm, s, ln, cat_col):
    if s >= 4.5:   glbl = "Very High"; gcol = "#16a34a"
    elif s >= 3.5: glbl = "High";      gcol = "#22c55e"
    else:          glbl = "Moderate";  gcol = "#4ade80"
    return (
        f"<div style='border-left:3px solid {cat_col};background:#1b222b;border-radius:8px;"
        f"padding:9px 12px;margin-bottom:9px'>"
        f"<b style='color:#e2e8f0'>{nm}</b> "
        f"<span style='background:{gcol};color:#021c0a;font-size:10px;font-weight:700;"
        f"padding:1px 7px;border-radius:10px'>{glbl}</span><br>"
        f"<span style='color:#94a3b8;font-size:12.5px;line-height:1.5'>{ln}</span></div>"
    )

def info_pill(label, value, color="#4a90a4"):
    if value in (None, ""): return ""
    return (
        f"<div style='background:#1b222b;border-radius:6px;padding:8px 12px;"
        f"margin-bottom:8px;border-left:2px solid {color}'>"
        f"<span style='color:{color};font-size:10px;font-weight:700;"
        f"text-transform:uppercase;letter-spacing:.5px'>{label}</span><br>"
        f"<span style='color:#d0d8e4;font-size:13px'>{value}</span></div>"
    )

# ── MAIN UI ───────────────────────────────────────────────────────────────────
st.title("Cnergee · Market Sales Briefing")
st.caption("Upload the matrix → select a country → see the levers to press.")

up = st.file_uploader("📂 Upload Marketing Matrix (.xlsx)", type=["xlsx"])
if not up:
    st.info("Upload your Marketing Matrix Excel file above to begin.")
    st.stop()

file_bytes = up.read()
headers, rows = load_sheet(file_bytes)
all_names = sorted(rows.keys())

col_s, col_m = st.columns([1, 3])
with col_s:
    q = st.text_input("🔍 Search country", "")
filtered = [n for n in all_names if q.lower() in n.lower()] if q else all_names
with col_m:
    picked = st.multiselect("Select country/countries", filtered,
                            default=filtered[:1] if filtered else [])

if not picked:
    st.info("Select one or more countries to see the briefing.")
    st.stop()

# ── PER-COUNTRY BRIEFING ──────────────────────────────────────────────────────
for name in picked:
    d = rows[name]
    st.markdown(f"---\n## {name}")

    # ─── 1. TAGLINES ───────────────────────────────────────────────────────
    tag_pos = d.get(COLS["tag_pos"])
    tag_neg = d.get(COLS["tag_neg"])
    if tag_pos or tag_neg:
        tc1, tc2 = st.columns(2)
        with tc1:
            if tag_pos:
                st.markdown(
                    f"<div style='background:linear-gradient(135deg,#0f2027,#203a43);"
                    f"border:1px solid #2dd4bf;border-radius:10px;padding:14px 18px;margin-bottom:14px'>"
                    f"<div style='color:#2dd4bf;font-size:9px;letter-spacing:2px;font-weight:700;"
                    f"text-transform:uppercase;margin-bottom:6px'>POSITIVE BRAND LINE</div>"
                    f"<div style='color:#fff;font-size:15px;font-weight:700;line-height:1.45'>{tag_pos}</div></div>",
                    unsafe_allow_html=True)
        with tc2:
            if tag_neg:
                st.markdown(
                    f"<div style='background:linear-gradient(135deg,#1a0a0a,#2d1515);"
                    f"border:1px solid #f97316;border-radius:10px;padding:14px 18px;margin-bottom:14px'>"
                    f"<div style='color:#f97316;font-size:9px;letter-spacing:2px;font-weight:700;"
                    f"text-transform:uppercase;margin-bottom:6px'>VALUE CHALLENGE LINE</div>"
                    f"<div style='color:#fff;font-size:15px;font-weight:700;line-height:1.45'>{tag_neg}</div></div>",
                    unsafe_allow_html=True)

    # ─── 2. KEY MARKET SNAPSHOT ───────────────────────────────────────────
    snap = st.columns(4)
    with snap[0]:
        gdp_v = d.get(COLS["gdp"])
        if gdp_v: st.metric("GDP", str(gdp_v))
    with snap[1]:
        gw = d.get(COLS["growth"])
        if gw: st.metric("GDP Growth", str(gw))
    with snap[2]:
        cur = d.get(COLS["curr_usd"])
        if cur: st.metric("vs USD", str(cur))
    with snap[3]:
        geo = d.get(COLS["geo_loc"])
        if geo: st.metric("Strategic Position", str(geo)[:40])

    # ─── 3. TAM / SAM / SOM ──────────────────────────────────────────────
    t_wifi = d.get(COLS["tam_wifi"]); t_sdwan = d.get(COLS["tam_sdwan"])
    t_ngfw = d.get(COLS["tam_ngfw"]); tot_tam = d.get(COLS["total_tam"])
    tot_sam = d.get(COLS["total_sam"]); tot_som = d.get(COLS["total_som"])
    if any(v not in (None, "") for v in [tot_tam, tot_sam, tot_som]):
        st.markdown("#### 📊 Market Opportunity (TAM / SAM / SOM)")
        tc = st.columns(6)
        for col_obj, lbl, val in zip(tc,
                                ["WiFi TAM","SD-WAN TAM","NGFW TAM","Total TAM","SAM","SOM"],
                                [t_wifi, t_sdwan, t_ngfw, tot_tam, tot_sam, tot_som]):
            with col_obj:
                display = f"${val}M" if val not in (None,"") and lbl != "Total TAM" else (str(val) if val else "—")
                if lbl in ("Total TAM","SAM","SOM"):
                    display = str(val) if val not in (None,"") else "—"
                st.metric(lbl, display)
    

    # ─── 4. COST ANIMATION ───────────────────────────────────────────────
    cust = parse_money(d.get(COLS["customer_tco"]))
    cnrg = parse_money(d.get(COLS["cnergee_cost"]))
    svgs = parse_money(d.get(COLS["savings_tco"]))
    st.markdown("#### 💸 Cost — What You're Bleeding vs Cnergee")
    if cust and cnrg and cust > cnrg:
        components.html(anim_html(cust, cnrg), height=235)
    else:
        st.markdown(
            "<div style='background:#1a1410;border:1px dashed #5a3a1a;border-radius:10px;"
            "padding:10px 14px;color:#9a8a72;font-size:12.5px'>"
            "TCO data not available — fill Customer TCO / Cnergee Cost columns and re-upload.</div>",
            unsafe_allow_html=True)

    # ─── 5. DEMOGRAPHICS ─────────────────────────────────────────────────
    lit = d.get(COLS["literacy"]); pu = d.get(COLS["pop_u"]); pr = d.get(COLS["pop_r"])
    if any(v not in (None,"") for v in [lit, pu, pr]):
        st.markdown("#### 👥 Demographics")
        dc = st.columns(3)
        with dc[0]:
            if lit: st.metric("Literacy Rate", f"{lit}%")
        with dc[1]:
            if pu: st.metric("Urban Population", f"{pu}%")
        with dc[2]:
            if pr: st.metric("Rural Population", f"{pr}%")

    # ─── 6. FEARS & GREEDS ───────────────────────────────────────────────
    st.markdown("---")
    fc_col, gc_col = st.columns(2)

    with fc_col:
        st.markdown("#### ⚠ Fears to Press On")
        F = fears(d)
        if not F:
            st.write("_No major risk flags for this market — lead on efficiency and cost._")
        for nm, lbl, cls, ln in F:
            st.markdown(fear_card(nm, lbl, cls, ln), unsafe_allow_html=True)

    with gc_col:
        st.markdown("#### ▲ Greed to Ride")
        G = greeds(d)
        if not G:
            st.write("_Stable / mature market — lead on reliability and predictable TCO._")
        for cat in GREED_CATS:
            if cat not in G: continue
            cat_col = GREED_COLS[cat]
            st.markdown(
                f"<div style='color:{cat_col};font-size:10px;font-weight:700;"
                f"letter-spacing:1.5px;text-transform:uppercase;margin:12px 0 5px;"
                f"border-bottom:1px solid {cat_col}44;padding-bottom:3px'>{cat}</div>",
                unsafe_allow_html=True)
            for s, nm, ln in G[cat]:
                st.markdown(greed_card(nm, s, ln, cat_col), unsafe_allow_html=True)

    # ─── 7. MARKET CONTEXT EXPANDER ──────────────────────────────────────
    with st.expander("📡 Market Context — Telecom, Pricing & Trade"):
        mc1, mc2 = st.columns(2)
        with mc1:
            blk = ""
            for lbl, key in [
                ("Top Telecom Operator", "top_telecom"), ("Local ISPs", "local_isp"),
                ("ISP Competition",      "isp_comp"),    ("Telecom Revenue Trend","telecom_trend"),
                ("Current ARPU (USD/mo)","arpu"),        ("Projected ARPU 2030", "arpu_2030"),
                ("Annual Telecom Churn", "churn"),       ("Managed Service Market","msp_opp"),
                ("Automation Spend",     "auto_spend"),  ("Projected Auto Spend", "auto_proj"),
            ]:
                v = d.get(COLS[key])
                if v not in (None,""):
                    val = f"{v}%" if key == "churn" and v else str(v)
                    blk += info_pill(lbl, val, "#0891b2")
            st.markdown(blk, unsafe_allow_html=True)
        with mc2:
            blk2 = ""
            for lbl, key in [
                ("Leased Line $/Mbps/mo",  "ill_price"),  ("4G/5G $/Mbps/mo",   "fg_price"),
                ("Broadband $/Mbps/mo",    "bb_price"),   ("Internet (Urban)",   "inet_u"),
                ("Internet (Rural)",       "inet_r"),     ("Cyber Spend % GDP",  "cyber_gdp"),
                ("Currency vs USD",        "curr_usd"),   ("Currency vs INR",    "curr_inr"),
                ("Data Sovereignty Laws",  "data_sov"),   ("Embargos/Sanctions", "embargo"),
            ]:
                v = d.get(COLS[key])
                if v not in (None,""):
                    blk2 += info_pill(lbl, str(v), "#7c3aed")
            st.markdown(blk2, unsafe_allow_html=True)

    # ─── 8. PSYCHOLOGY & CULTURE ─────────────────────────────────────────
    psych_fields = [
        ("Human Psychology & Negotiation", "psych"),
        ("Brand Consciousness",            "brand"),
        ("Colors & Visual Culture",        "colors"),
        ("Religion & Beliefs",             "belief"),
    ]
    psych_rows = [(lbl, str(d.get(COLS[k]))) for lbl, k in psych_fields
                  if d.get(COLS[k]) not in (None, "")]
    if psych_rows:
        st.markdown("#### 🧠 Psychology & Culture")
        body = "".join(
            f"<div style='margin-bottom:7px'>"
            f"<b style='color:#a78bfa'>{lbl}:</b> "
            f"<span style='color:#cbd5e1;font-size:13px'>{val}</span></div>"
            for lbl, val in psych_rows
        )
        st.markdown(
            f"<div style='background:#1b222b;border-left:3px solid #8b5cf6;"
            f"border-radius:8px;padding:12px 15px'>{body}</div>",
            unsafe_allow_html=True)

    # ─── 9. EMBARGO BANNER ───────────────────────────────────────────────
    emb = d.get(COLS["embargo"])
    if emb not in (None, ""):
        clear = str(emb).strip().lower() in ("none","no","nil","—","-","n/a")
        ec = "#3a4654" if clear else "#ef4444"
        st.markdown(
            f"<div style='border-left:4px solid {ec};background:#1b222b;border-radius:8px;"
            f"padding:10px 14px;margin:8px 0'>"
            f"<span style='color:{ec};font-size:11px;text-transform:uppercase;"
            f"letter-spacing:.5px;font-weight:700'>⚑ Embargos / Sanctions</span><br>"
            f"<span style='font-size:13.5px;color:#e2e8f0'>{emb}</span></div>",
            unsafe_allow_html=True)

    # ─── 10. DEEP LOOK ────────────────────────────────────────────────────
    with st.expander("🔍 Deep Look — every selected column for this country"):
        table = []
        for key, cn in COLS.items():
            v = d.get(cn)
            if v not in (None, ""):
                table.append({
                    "Col #": cn,
                    "Field": headers.get(cn, f"Col {cn}"),
                    "Key":   key,
                    "Value": str(v),
                })
        if table:
            st.dataframe(pd.DataFrame(table), hide_index=True, use_container_width=True)
        else:
            st.write("No data found for this country.")